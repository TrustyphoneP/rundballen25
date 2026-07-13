"""
apps/rechnungen/views.py – Rechnungsführung

Belege hochladen, automatisch analysieren, Positionen prüfen und
Grundpreise in die Zutaten-Stammdaten übernehmen. Darauf aufbauend:
Preisliste, Rezeptkosten, Freizeitkosten und Preistreiber.
"""
import logging
from decimal import Decimal, InvalidOperation

import datetime

from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.text import slugify
from django.views.decorators.http import require_POST

from apps.camps.models import Camp
from apps.recipes.models import Ingredient

from . import services
from .forms import BelegUploadForm, KostenkategorieForm
from .models import Beleg, BelegPosition, Kostenkategorie, Preishistorie

logger = logging.getLogger(__name__)


def _aktive_freizeit():
    return Camp.objects.filter(is_active=True).order_by("-start_date").first()


# ---------------------------------------------------------------------------
# Belegübersicht
# ---------------------------------------------------------------------------

@login_required
def index(request):
    camp = _aktive_freizeit()
    belege = []
    ist_summe = None
    ohne_betrag = 0
    if camp:
        belege = (
            Beleg.objects.filter(camp=camp)
            .select_related("hochgeladen_von", "kategorie")
            .annotate(anzahl_positionen=Count("positionen"))
        )
        ist_summe = belege.filter(nur_preiserfassung=False).aggregate(s=Sum("gesamtbetrag"))["s"]
        # Belege ohne Gesamtbetrag (neu oder Fehler) fehlen still in der
        # Ist-Summe -- das muss sichtbar sein, sonst täuscht die Zahl.
        ohne_betrag = Beleg.objects.filter(
            camp=camp, nur_preiserfassung=False, gesamtbetrag__isnull=True
        ).count()

    return render(request, "rechnungen/index.html", {
        "camp":        camp,
        "belege":      belege,
        "ist_summe":   ist_summe,
        "ohne_betrag": ohne_betrag,
    })


# ---------------------------------------------------------------------------
# Beleg hochladen
# ---------------------------------------------------------------------------

@login_required
def upload(request):
    camp = _aktive_freizeit()
    if not camp:
        messages.warning(request, "Keine aktive Freizeit vorhanden -- bitte zuerst eine Freizeit aktivieren.")
        return redirect("rechnungen:index")

    form = BelegUploadForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        bilder = request.FILES.getlist("bilder")
        kategorie = form.cleaned_data["kategorie"]
        nur_preiserfassung = form.cleaned_data["nur_preiserfassung"]
        sofort = form.cleaned_data["sofort_analysieren"]
        notiz = form.cleaned_data["notiz"]

        neue_belege = []
        for bild in bilder:
            beleg = Beleg.objects.create(
                camp=camp,
                kategorie=kategorie,
                nur_preiserfassung=nur_preiserfassung,
                image=bild,
                notiz=notiz,
                hochgeladen_von=request.user,
            )
            neue_belege.append(beleg)

        if sofort:
            fehler = 0
            for beleg in neue_belege:
                erfolg, meldung = services.analysiere_beleg(beleg, user=request.user)
                if not erfolg:
                    fehler += 1
                    messages.error(request, f"Beleg {beleg.pk}: {meldung}")
            ok = len(neue_belege) - fehler
            if ok:
                messages.success(request, f"{ok} Beleg(e) hochgeladen und analysiert.")
        else:
            messages.success(request, f"{len(neue_belege)} Beleg(e) hochgeladen. Analyse kann pro Beleg gestartet werden.")

        if len(neue_belege) == 1:
            return redirect("rechnungen:detail", pk=neue_belege[0].pk)
        return redirect("rechnungen:index")

    return render(request, "rechnungen/upload.html", {"form": form, "camp": camp})


# ---------------------------------------------------------------------------
# Beleg-Detail: Bild, Positionen, Zuordnung, Preisübernahme
# ---------------------------------------------------------------------------

@login_required
def detail(request, pk):
    beleg = get_object_or_404(
        Beleg.objects.select_related("camp").prefetch_related("positionen__ingredient"),
        pk=pk,
    )
    zutaten_namen = list(Ingredient.objects.order_by("name").values_list("name", flat=True))
    positionen = list(beleg.positionen.all())
    differenz = None
    if beleg.gesamtbetrag is not None:
        differenz = beleg.gesamtbetrag - beleg.positionen_summe

    # Überschreiben-Rückfrage: Bei reinen Preiserfassungs-Belegen (alte
    # Bons) pro Position prüfen, ob die Zutat bereits einen ABWEICHENDEN
    # Preis hat. Dann fragt der Übernehmen-Button per Popup nach, bevor
    # der alte Beleg den aktuellen Preis überschreibt.
    abweichende_anzahl = 0
    for pos in positionen:
        pos.preis_weicht_ab = False
        pos.neuer_preis = None
        pos.alter_preis = None
        pos.vergleichs_einheit = ""
        if not beleg.nur_preiserfassung or pos.ingredient is None or pos.ist_uebernommen:
            continue
        neuer, einheit, fehler = services.preis_fuer_uebernahme(pos)
        if fehler or neuer is None:
            continue
        alter = services.aktueller_preis_in(pos.ingredient, einheit)
        if alter is not None and alter != neuer:
            pos.preis_weicht_ab = True
            pos.neuer_preis = neuer
            pos.alter_preis = alter
            pos.vergleichs_einheit = einheit
            abweichende_anzahl += 1

    return render(request, "rechnungen/detail.html", {
        "beleg":         beleg,
        "positionen":    positionen,
        "zutaten_namen": zutaten_namen,
        "differenz":     differenz,
        "einheiten":     Ingredient.Unit.choices,
        "grundpreis_einheiten": BelegPosition.GrundpreisEinheit.choices,
        "kategorien":    Kostenkategorie.objects.all(),
        "abweichende_anzahl": abweichende_anzahl,
    })


@login_required
@require_POST
def beleg_kategorie(request, pk):
    """Ändert die Kostenkategorie eines Belegs."""
    beleg = get_object_or_404(Beleg, pk=pk)
    kategorie_id = request.POST.get("kategorie")
    if kategorie_id:
        kategorie = get_object_or_404(Kostenkategorie, pk=kategorie_id)
        beleg.kategorie = kategorie
        beleg.save(update_fields=["kategorie"])
        messages.success(request, f"Kategorie auf '{kategorie.name}' gesetzt.")
    else:
        beleg.kategorie = None
        beleg.save(update_fields=["kategorie"])
        messages.success(request, "Kategorie entfernt.")
    return redirect("rechnungen:detail", pk=beleg.pk)


@login_required
@require_POST
def analysieren(request, pk):
    beleg = get_object_or_404(Beleg, pk=pk)
    erfolg, meldung = services.analysiere_beleg(beleg, user=request.user)
    if erfolg:
        messages.success(request, meldung)
    else:
        messages.error(request, meldung)
    return redirect("rechnungen:detail", pk=beleg.pk)


@login_required
@require_POST
def loeschen(request, pk):
    beleg = get_object_or_404(Beleg, pk=pk)
    beleg.image.delete(save=False)
    beleg.delete()
    messages.success(request, "Beleg gelöscht.")
    return redirect("rechnungen:index")


def _parse_decimal_de(wert, nachkommastellen="0.0001"):
    """Formulareingabe (deutsches Komma erlaubt) -> Decimal oder None."""
    wert = (wert or "").strip().replace(",", ".")
    if not wert:
        return None
    try:
        return Decimal(wert).quantize(Decimal(nachkommastellen))
    except InvalidOperation:
        return None


@login_required
@require_POST
def position_aktualisieren(request, pk):
    """
    Speichert die manuell geprüften Felder einer Position. Bei
    action=uebernehmen wird zusätzlich der Grundpreis in die
    Zutaten-Stammdaten übertragen.
    """
    position = get_object_or_404(
        BelegPosition.objects.select_related("beleg", "ingredient"), pk=pk
    )

    position.artikel_name = (request.POST.get("artikel_name") or position.artikel_name)[:200]
    position.menge = _parse_decimal_de(request.POST.get("menge"), "0.001")

    einheit = request.POST.get("einheit") or ""
    gueltige_einheiten = {code for code, _ in Ingredient.Unit.choices}
    position.einheit = einheit if einheit in gueltige_einheiten else ""

    position.gesamtpreis = _parse_decimal_de(request.POST.get("gesamtpreis"), "0.01")
    position.grundpreis = _parse_decimal_de(request.POST.get("grundpreis"))

    gp_einheit = request.POST.get("grundpreis_einheit") or ""
    gueltige_gp = {code for code, _ in BelegPosition.GrundpreisEinheit.choices}
    position.grundpreis_einheit = gp_einheit if gp_einheit in gueltige_gp else ""

    position.stueckpreis = _parse_decimal_de(request.POST.get("stueckpreis"))

    # Grundpreis ggf. aus Menge + Zeilensumme nachrechnen
    if position.grundpreis is None or not position.grundpreis_einheit:
        gp, gpe = services.ermittle_grundpreis(
            position.menge, position.einheit, position.gesamtpreis
        )
        if gp is not None:
            position.grundpreis, position.grundpreis_einheit = gp, gpe

    # Zutat per Name auflösen (Datalist-Eingabe), leer = keine Zuordnung
    zutat_name = (request.POST.get("zutat") or "").strip()
    if zutat_name:
        ing = services.lookup_ingredient(zutat_name)
        if ing is None:
            messages.warning(
                request,
                f"Zutat '{zutat_name}' nicht in den Stammdaten gefunden -- "
                "Zuordnung unverändert gelassen.",
            )
        else:
            if ing != position.ingredient:
                position.ist_uebernommen = False
                position.uebernommen_am = None
            position.ingredient = ing
    else:
        position.ingredient = None
        position.ist_uebernommen = False
        position.uebernommen_am = None

    position.save()

    if request.POST.get("action") == "uebernehmen":
        erfolg, meldung = services.uebernehme_position(position, user=request.user)
        if erfolg:
            messages.success(request, meldung)
        else:
            messages.error(request, meldung)
    else:
        messages.success(request, f"Position '{position.artikel_name}' gespeichert.")

    if position.beleg.status == Beleg.Status.GEPRUEFT:
        position.beleg.status = Beleg.Status.ANALYSIERT
        position.beleg.save(update_fields=["status"])
        messages.info(request, "Beleg-Prüfung zurückgesetzt, da Positionen geändert wurden.")

    return redirect("rechnungen:detail", pk=position.beleg.pk)


@login_required
@require_POST
def position_loeschen(request, pk):
    position = get_object_or_404(BelegPosition.objects.select_related("beleg"), pk=pk)
    beleg = position.beleg
    position.delete()
    messages.success(request, "Position gelöscht.")
    if beleg.status == Beleg.Status.GEPRUEFT:
        beleg.status = Beleg.Status.ANALYSIERT
        beleg.save(update_fields=["status"])
        messages.info(request, "Beleg-Prüfung zurückgesetzt, da Positionen geändert wurden.")
    return redirect("rechnungen:detail", pk=beleg.pk)


@login_required
@require_POST
def alle_uebernehmen(request, pk):
    """Überträgt alle zugeordneten, noch nicht übernommenen Grundpreise."""
    beleg = get_object_or_404(Beleg, pk=pk)
    kandidaten = beleg.positionen.filter(
        ingredient__isnull=False, ist_uebernommen=False,
        grundpreis__isnull=False,
    ).exclude(grundpreis_einheit="")

    ok, fehler = 0, 0
    for position in kandidaten:
        erfolg, meldung = services.uebernehme_position(position, user=request.user)
        if erfolg:
            ok += 1
        else:
            fehler += 1
            messages.error(request, meldung)

    if ok:
        messages.success(request, f"{ok} Preis(e) in die Zutaten-Stammdaten übernommen.")
    if not ok and not fehler:
        messages.info(request, "Keine offenen, zugeordneten Positionen mit Grundpreis vorhanden.")
    return redirect("rechnungen:detail", pk=beleg.pk)


@login_required
@require_POST
def pruefen(request, pk):
    """
    Markiert einen analysierten Beleg als geprüft (Positionen kontrolliert,
    Summe stimmt) bzw. nimmt die Prüfung wieder zurück. Teil des
    Dokumentations-Workflows: neu -> analysiert -> geprüft.
    """
    beleg = get_object_or_404(Beleg, pk=pk)
    if beleg.status == Beleg.Status.GEPRUEFT:
        beleg.status = Beleg.Status.ANALYSIERT
        beleg.save(update_fields=["status"])
        messages.info(request, "Prüfung zurückgenommen.")
    elif beleg.status == Beleg.Status.ANALYSIERT:
        beleg.status = Beleg.Status.GEPRUEFT
        beleg.save(update_fields=["status"])
        messages.success(request, "Beleg als geprüft markiert.")
    else:
        messages.error(request, "Nur analysierte Belege können als geprüft markiert werden.")
    return redirect("rechnungen:detail", pk=beleg.pk)


# ---------------------------------------------------------------------------
# Excel-Export der Belegliste (Ausgabendokumentation)
# ---------------------------------------------------------------------------

@login_required
def export_belege(request):
    """
    Belegliste der aktiven Freizeit als Excel-Datei: ein Blatt mit allen
    Belegen (ohne reine Preiserfassungs-Belege), ein Blatt mit Summen je
    Kategorie. Das ist das Endprodukt der Ausgabendokumentation, z.B. für
    die Abrechnung gegenüber dem Träger.
    """
    camp = _aktive_freizeit()
    if not camp:
        messages.warning(request, "Keine aktive Freizeit vorhanden.")
        return redirect("rechnungen:index")

    import openpyxl
    from io import BytesIO
    from openpyxl.styles import Font
    from openpyxl.utils import get_column_letter

    basis = Beleg.objects.filter(camp=camp, nur_preiserfassung=False)
    belege = (
        basis.select_related("kategorie", "hochgeladen_von")
        .annotate(anzahl_positionen=Count("positionen"))
        .order_by("kaufdatum", "pk")
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Belege"
    ws.append([
        "Kaufdatum", "Händler", "Kategorie", "Betrag (€)",
        "Status", "Geprüft", "Positionen", "Hochgeladen von", "Notiz",
    ])
    for zelle in ws[1]:
        zelle.font = Font(bold=True)

    for beleg in belege:
        hochgeladen_von = ""
        if beleg.hochgeladen_von:
            hochgeladen_von = (
                beleg.hochgeladen_von.get_full_name()
                or beleg.hochgeladen_von.username
            )
        ws.append([
            beleg.kaufdatum,
            beleg.haendler,
            beleg.kategorie.name if beleg.kategorie else "",
            float(beleg.gesamtbetrag) if beleg.gesamtbetrag is not None else None,
            beleg.get_status_display(),
            "Ja" if beleg.status == Beleg.Status.GEPRUEFT else "Nein",
            beleg.anzahl_positionen,
            hochgeladen_von,
            beleg.notiz,
        ])

    for zeile in ws.iter_rows(min_row=2, min_col=1, max_col=1):
        for zelle in zeile:
            zelle.number_format = "DD.MM.YYYY"
    for zeile in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for zelle in zeile:
            zelle.number_format = '#,##0.00" €"'
    for spalte, breite in enumerate([12, 26, 22, 12, 18, 9, 11, 20, 32], start=1):
        ws.column_dimensions[get_column_letter(spalte)].width = breite

    # Blatt 2: Summen je Kategorie
    ws2 = wb.create_sheet("Summen je Kategorie")
    ws2.append(["Kategorie", "Belege", "Summe (€)"])
    for zelle in ws2[1]:
        zelle.font = Font(bold=True)

    gesamt = Decimal("0")
    gesamt_anzahl = 0
    for kategorie in Kostenkategorie.objects.all():
        qs = basis.filter(kategorie=kategorie)
        anzahl = qs.count()
        if anzahl == 0:
            continue
        summe = qs.aggregate(s=Sum("gesamtbetrag"))["s"] or Decimal("0")
        ws2.append([kategorie.name, anzahl, float(summe)])
        gesamt += summe
        gesamt_anzahl += anzahl

    ohne_qs = basis.filter(kategorie__isnull=True)
    ohne_anzahl = ohne_qs.count()
    if ohne_anzahl:
        ohne_summe = ohne_qs.aggregate(s=Sum("gesamtbetrag"))["s"] or Decimal("0")
        ws2.append(["Ohne Kategorie", ohne_anzahl, float(ohne_summe)])
        gesamt += ohne_summe
        gesamt_anzahl += ohne_anzahl

    ws2.append(["Gesamt", gesamt_anzahl, float(gesamt)])
    for zelle in ws2[ws2.max_row]:
        zelle.font = Font(bold=True)

    ohne_betrag = basis.filter(gesamtbetrag__isnull=True).count()
    if ohne_betrag:
        ws2.append([])
        ws2.append([f"Hinweis: {ohne_betrag} Beleg(e) ohne Gesamtbetrag "
                    "(neu oder Fehler) -- Summen unvollständig."])

    for zeile in ws2.iter_rows(min_row=2, min_col=3, max_col=3):
        for zelle in zeile:
            zelle.number_format = '#,##0.00" €"'
    for spalte, breite in enumerate([26, 10, 14], start=1):
        ws2.column_dimensions[get_column_letter(spalte)].width = breite

    buf = BytesIO()
    wb.save(buf)
    dateiname = f"belege_{slugify(camp.name)}_{datetime.date.today().isoformat()}.xlsx"
    response = HttpResponse(
        buf.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{dateiname}"'
    return response


# ---------------------------------------------------------------------------
# Kategoriekosten: Ist-Ausgaben je Kostenkategorie, Kategorien verwalten
# ---------------------------------------------------------------------------

@login_required
def kategorien(request):
    """
    Zwei getrennte Sichten auf die Kosten der aktiven Freizeit:

    1. Ist-Kosten je Kostenkategorie -- Summe der Beleg-Gesamtbeträge,
       gruppiert nach der (optionalen) Kategorie am Beleg. Das sind die
       tatsächlichen Ausgaben und die führende Größe.
    2. Vorkalkulation je Einkaufslisten-Kategorie -- Zutatenpreise mal
       Mengen aus build_shopping_day_items, gruppiert nach der Quelle der
       Einkaufsliste (Abendessen, Frühstück/Mittag, Allgemein,
       Betreueressen, Alternative). Reine Planungsübersicht.

    Dazu Verwaltung: neue Kostenpunkte anlegen, unbenutzte löschen.
    """
    camp = _aktive_freizeit()

    form = KostenkategorieForm(request.POST or None)
    if request.method == "POST" and form.is_valid():
        kategorie = form.save(commit=False)
        max_sort = Kostenkategorie.objects.order_by("-sortierung").first()
        kategorie.sortierung = (max_sort.sortierung + 10) if max_sort else 10
        kategorie.save()
        messages.success(request, f"Kostenkategorie '{kategorie.name}' angelegt.")
        return redirect("rechnungen:kategorien")

    # --- 1. Ist-Kosten aus Belegen ---
    eintraege = []
    ist_gesamt = Decimal("0")
    for kategorie in Kostenkategorie.objects.all():
        belege = kategorie.belege.filter(camp=camp, nur_preiserfassung=False) if camp else kategorie.belege.none()
        summe = belege.aggregate(s=Sum("gesamtbetrag"))["s"] or Decimal("0")
        anzahl = belege.count()
        ist_gesamt += summe
        eintraege.append({
            "kategorie": kategorie,
            "summe":     summe,
            "anzahl":    anzahl,
            "loeschbar": not kategorie.belege.exists(),
        })

    ohne = Beleg.objects.filter(camp=camp, kategorie__isnull=True, nur_preiserfassung=False) if camp else Beleg.objects.none()
    ohne_summe = ohne.aggregate(s=Sum("gesamtbetrag"))["s"] or Decimal("0")
    ist_gesamt += ohne_summe

    ohne_betrag = 0
    if camp:
        ohne_betrag = Beleg.objects.filter(
            camp=camp, nur_preiserfassung=False, gesamtbetrag__isnull=True
        ).count()

    # --- 2. Vorkalkulation aus der Einkaufsliste (Zutatenpreise) ---
    plan_quellen = []
    plan_summe = Decimal("0")
    plan_ohne_preis = []
    if camp:
        plan = services.berechne_freizeitkosten(camp)
        plan_quellen = plan["quellen"]
        plan_summe = plan["summe"]
        plan_ohne_preis = plan["ohne_preis"]

    return render(request, "rechnungen/kategorien.html", {
        "camp":            camp,
        "eintraege":       eintraege,
        "ohne_anzahl":     ohne.count(),
        "ohne_summe":      ohne_summe,
        "ohne_betrag":     ohne_betrag,
        "ist_gesamt":      ist_gesamt,
        "plan_quellen":    plan_quellen,
        "plan_summe":      plan_summe,
        "plan_ohne_preis": plan_ohne_preis,
        "form":            form,
    })


@login_required
@require_POST
def kategorie_loeschen(request, pk):
    kategorie = get_object_or_404(Kostenkategorie, pk=pk)
    if kategorie.belege.exists():
        messages.error(
            request,
            f"'{kategorie.name}' hat zugeordnete Belege und kann nicht gelöscht werden.",
        )
    else:
        name = kategorie.name
        kategorie.delete()
        messages.success(request, f"Kostenkategorie '{name}' gelöscht.")
    return redirect("rechnungen:kategorien")


# ---------------------------------------------------------------------------
# Preisliste
# ---------------------------------------------------------------------------

@login_required
def preisliste(request):
    q = request.GET.get("q", "").strip()
    nur_ohne = request.GET.get("ohne") == "1"

    zutaten = Ingredient.objects.all().prefetch_related("preishistorie")
    if q:
        zutaten = zutaten.filter(name__icontains=q)
    if nur_ohne:
        zutaten = zutaten.filter(price__isnull=True)

    eintraege = []
    mit_preis = 0
    for ing in zutaten:
        # Prefetch nutzen statt zu slicen -- .all()[:1] würde pro Zutat
        # eine neue Query auslösen. Historie ist per Meta absteigend sortiert.
        historie = list(ing.preishistorie.all())
        letzter = historie[0] if historie else None
        if ing.price is not None:
            mit_preis += 1
        eintraege.append({
            "ingredient": ing,
            "letzter":    letzter,
            "anzahl":     len(historie),
        })

    return render(request, "rechnungen/preisliste.html", {
        "eintraege": eintraege,
        "q":         q,
        "nur_ohne":  nur_ohne,
        "mit_preis": mit_preis,
        "gesamt":    len(eintraege),
    })


# ---------------------------------------------------------------------------
# Rezeptkosten
# ---------------------------------------------------------------------------

@login_required
def rezeptkosten(request):
    ergebnisse = services.berechne_rezeptkosten()
    return render(request, "rechnungen/rezeptkosten.html", {
        "ergebnisse": ergebnisse,
    })


# ---------------------------------------------------------------------------
# Freizeitkosten
# ---------------------------------------------------------------------------

@login_required
def freizeitkosten(request):
    camp = _aktive_freizeit()
    if not camp:
        messages.warning(request, "Keine aktive Freizeit vorhanden.")
        return redirect("rechnungen:index")

    kosten = services.berechne_freizeitkosten(camp)

    personen = camp.participants.count() or camp.total_persons
    pro_person = kosten["summe"] / personen if personen else None
    tage = camp.duration_days
    pro_person_tag = pro_person / tage if pro_person is not None and tage else None

    ist_summe = Beleg.objects.filter(camp=camp, nur_preiserfassung=False).aggregate(s=Sum("gesamtbetrag"))["s"]
    ohne_betrag = Beleg.objects.filter(
        camp=camp, nur_preiserfassung=False, gesamtbetrag__isnull=True
    ).count()

    return render(request, "rechnungen/freizeitkosten.html", {
        "camp":           camp,
        "kosten":         kosten,
        "personen":       personen,
        "tage":           tage,
        "pro_person":     pro_person,
        "pro_person_tag": pro_person_tag,
        "ist_summe":      ist_summe,
        "ohne_betrag":    ohne_betrag,
    })


# ---------------------------------------------------------------------------
# Zutaten-Steckbrief: alles Relevante zu einer Zutat auf einer Seite
# ---------------------------------------------------------------------------

@login_required
def zutat(request, pk):
    """
    Steckbrief einer Zutat: Preis je Einheit (inkl. Normalisierung auf
    €/kg bzw. €/l), tatsächliche Ausgaben aus Beleg-Positionen,
    Verwendung in Rezepten und in Allgemein/Betreueressen/Alternative,
    geplante Kosten in der aktiven Freizeit sowie die Preishistorie.
    """
    from apps.meals.models import GeneralIngredient
    from apps.recipes.models import RecipeIngredient

    ing = get_object_or_404(
        Ingredient.objects.prefetch_related("allergens"), pk=pk
    )
    camp = _aktive_freizeit()

    # Preis normalisiert auf die gebräuchliche Grundpreis-Einheit
    preis_normalisiert = None
    norm_einheit = ""
    if ing.price is not None and ing.price_unit:
        if ing.price_unit in ("g", "kg"):
            preis_normalisiert = services.konvertiere_preis(ing.price, ing.price_unit, "kg")
            norm_einheit = "kg"
        elif ing.price_unit in ("ml", "l"):
            preis_normalisiert = services.konvertiere_preis(ing.price, ing.price_unit, "l")
            norm_einheit = "l"

    # Verwendung in Rezepten (mit Kosten je Rezept-Ansatz und pro Person)
    rezept_verwendungen = []
    for ri in RecipeIngredient.objects.filter(ingredient=ing).select_related("recipe"):
        kosten = ing.price_for(ri.amount, ri.unit)
        pro_person = None
        if kosten is not None and ri.recipe.base_servings:
            pro_person = kosten / Decimal(str(ri.recipe.base_servings))
        rezept_verwendungen.append({
            "ri": ri, "kosten": kosten, "pro_person": pro_person,
        })

    # Verwendung in Allgemein / Betreueressen / SKF-Alternativen
    general_verwendungen = (
        GeneralIngredient.objects.filter(ingredient=ing)
        .select_related("camp").order_by("-camp__start_date", "category")
    )

    # Tatsächliche Ausgaben: Beleg-Positionen dieser Zutat
    positionen = (
        BelegPosition.objects.filter(ingredient=ing)
        .select_related("beleg", "beleg__kategorie")
        .order_by("-beleg__kaufdatum", "-beleg__hochgeladen_am")
    )
    ist_summe = Decimal("0")
    ist_anzahl = 0
    if camp:
        for pos in positionen:
            if (pos.beleg.camp_id == camp.pk
                    and not pos.beleg.nur_preiserfassung
                    and pos.gesamtpreis is not None):
                ist_summe += pos.gesamtpreis
                ist_anzahl += 1

    # Geplante Kosten in der aktiven Freizeit (aus der Einkaufsliste)
    plan_eintraege = []
    plan_summe = Decimal("0")
    if camp:
        kosten = services.berechne_freizeitkosten(camp)
        for lieferung in kosten["lieferungen"]:
            for p in lieferung["positionen"]:
                if p["name"] == ing.name:
                    plan_eintraege.append({
                        "lieferung": lieferung["label"],
                        "amount":    p["amount"],
                        "unit":      p["unit"],
                        "source":    p["source"],
                        "kosten":    p["kosten"],
                    })
                    if p["kosten"] is not None:
                        plan_summe += p["kosten"]

    historie = ing.preishistorie.select_related("beleg")[:50]

    return render(request, "rechnungen/zutat_detail.html", {
        "ing":                 ing,
        "camp":                camp,
        "einheiten":           ing.all_units_used(),
        "preis_normalisiert":  preis_normalisiert,
        "norm_einheit":        norm_einheit,
        "rezept_verwendungen": rezept_verwendungen,
        "general_verwendungen": general_verwendungen,
        "positionen":          positionen,
        "ist_summe":           ist_summe,
        "ist_anzahl":          ist_anzahl,
        "plan_eintraege":      plan_eintraege,
        "plan_summe":          plan_summe,
        "historie":            historie,
    })


# ---------------------------------------------------------------------------
# Preistreiber
# ---------------------------------------------------------------------------

@login_required
def preistreiber(request):
    treiber = services.berechne_preistreiber()

    camp = _aktive_freizeit()
    top_positionen = []
    if camp:
        top_positionen = services.berechne_freizeitkosten(camp)["top_positionen"]

    return render(request, "rechnungen/preistreiber.html", {
        "treiber":        treiber,
        "camp":           camp,
        "top_positionen": top_positionen,
    })
